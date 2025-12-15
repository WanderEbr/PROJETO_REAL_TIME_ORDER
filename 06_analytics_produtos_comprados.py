import csv
import os
from collections import Counter
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

# Diretório e arquivos
diretorio = r"C:\Users\wande\OneDrive\Documentos\_01_projeto_databricks_azure\archive"

arquivo_csv = os.path.join(
    diretorio,
    "sales_transaction_v_4a_tratado_data.csv"
)

arquivo_excel = os.path.join(
    diretorio,
    "total_produtos_comprados.xlsx"
)

# Contador de produtos
contador_produtos = Counter()

# ===== Leitura do CSV =====
with open(arquivo_csv, mode="r", encoding="utf-8", newline="") as arquivo:
    leitor = csv.reader(arquivo, delimiter=";")

    cabecalho = next(leitor)

    if "ProductName" not in cabecalho:
        raise Exception("Coluna 'ProductName' não encontrada no arquivo.")

    idx_produto = cabecalho.index("ProductName")

    # Dados a partir da linha 2
    for linha in leitor:
        if not linha or len(linha) <= idx_produto:
            continue

        produto = linha[idx_produto].strip()

        if produto:
            contador_produtos[produto] += 1

# Selecionar Top 10 produtos
top_10_produtos = contador_produtos.most_common(10)

# ===== Criação do Excel =====
wb = Workbook()
ws = wb.active
ws.title = "Top_Produtos"

# Cabeçalho
ws.append(["Produto", "Total de Compras"])

# Dados
for produto, total in top_10_produtos:
    ws.append([produto, total])

# ===== Gráfico de Barras =====
grafico = BarChart()
grafico.title = "Top 10 Produtos Mais Comprados"
grafico.y_axis.title = "Total de Compras"
grafico.x_axis.title = "Produto"

dados = Reference(
    ws,
    min_col=2,
    min_row=1,
    max_row=ws.max_row
)

categorias = Reference(
    ws,
    min_col=1,
    min_row=2,
    max_row=ws.max_row
)

grafico.add_data(dados, titles_from_data=True)
grafico.set_categories(categorias)

ws.add_chart(grafico, "D2")

# Salvar Excel
wb.save(arquivo_excel)

print("Arquivo Excel gerado com sucesso:")
print(arquivo_excel)
