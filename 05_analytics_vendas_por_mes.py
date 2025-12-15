import csv
import os
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference

# Diretório e arquivos
diretorio = r"C:\Users\wande\OneDrive\Documentos\_01_projeto_databricks_azure\archive"

arquivo_csv = os.path.join(
    diretorio,
    "sales_transaction_v_4a_tratado_data.csv"
)

arquivo_excel = os.path.join(
    diretorio,
    "total_vendas.xlsx"
)

# Acumulador de vendas por mês
vendas_por_mes = defaultdict(int)

def extrair_ano_mes(data_str):
    try:
        data = datetime.strptime(data_str.strip(), "%d/%m/%Y")
        return f"{data.year}-{data.month:02d}"
    except ValueError:
        return None

# ===== Leitura do CSV =====
with open(arquivo_csv, mode="r", encoding="utf-8", newline="") as arquivo:
    leitor = csv.reader(arquivo, delimiter=";")
    cabecalho = next(leitor)

    if "Date" not in cabecalho:
        raise Exception("Coluna 'Date' não encontrada.")

    idx_date = cabecalho.index("Date")

    for linha in leitor:
        if not linha or len(linha) <= idx_date:
            continue

        ano_mes = extrair_ano_mes(linha[idx_date])
        if ano_mes:
            vendas_por_mes[ano_mes] += 1

# ===== Criação do Excel =====
wb = Workbook()

# Aba 1 – Total de vendas mensal
ws_dados = wb.active
ws_dados.title = "Total_Vendas_Mensal"

ws_dados.append(["Ano-Mês", "Total de Vendas"])

for mes in sorted(vendas_por_mes):
    ws_dados.append([mes, vendas_por_mes[mes]])

# ===== Gráfico de Linha =====
grafico = LineChart()
grafico.title = "Tendência de Vendas Mensais"
grafico.y_axis.title = "Total de Vendas"
grafico.x_axis.title = "Ano-Mês"

dados = Reference(
    ws_dados,
    min_col=2,
    min_row=1,
    max_row=ws_dados.max_row
)

categorias = Reference(
    ws_dados,
    min_col=1,
    min_row=2,
    max_row=ws_dados.max_row
)

grafico.add_data(dados, titles_from_data=True)
grafico.set_categories(categorias)

ws_dados.add_chart(grafico, "D2")

# ===== Aba 2 – Análise de Tendência =====
ws_analise = wb.create_sheet(title="Analise_Tendencia")

ws_analise["A1"] = "Análise de Tendência de Vendas"
ws_analise["A3"] = (
    "A análise das vendas mês a mês indica um comportamento sazonal, "
    "com picos em determinados períodos e quedas em outros, "
    "não caracterizando um crescimento linear contínuo. "
    "O volume de vendas varia significativamente ao longo do tempo, "
    "o que sugere influência de campanhas, eventos ou sazonalidade do negócio."
)

# Salvar arquivo
wb.save(arquivo_excel)

print("Arquivo Excel com análise e gráfico gerado com sucesso:")
print(arquivo_excel)
